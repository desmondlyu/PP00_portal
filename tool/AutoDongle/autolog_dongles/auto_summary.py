from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import re
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook


@dataclass(frozen=True)
class PortData:
    name: str
    lines: List[str]


@dataclass(frozen=True)
class SummaryResult:
    output_path: Path
    warnings: List[str]


_LABEL_AVG = "PGM Costtimes(Average)(us):"
_LABEL_MAX = "PGM Costtimes(MAX)(us):"
_LABEL_TBE = "Erase Costtimes(Block)(ms):"
_LABEL_TSE = "Erase Costtimes(Sector)(ms):"
_LABEL_TBE_ALT = "Erase Costtimes(ms):"
_PASS_LINE = "= = = = = = Result : PASS !! = = = = = ="
_LOG_END_LINE = "= = = = = = = = LOG END = = = = = = = ="
# 支援 Socket's UID 與 Socket Flash UID，允許不同空白與分隔符
_UID_PATTERN = re.compile(
    r"Socket(?:'s)?(?:\s+Flash)?\s+UID\s*(?:[:;]\s*)+([0-9A-Fa-f]+)",
    re.IGNORECASE,
)


def run_auto_summary(log_dir: Path, cyc_type: int) -> SummaryResult:
    if not log_dir.exists():
        return SummaryResult(output_path=_build_output_path(log_dir), warnings=["empty_log"])

    port_entries = _load_port_data(log_dir)
    if not port_entries:
        return SummaryResult(output_path=_build_output_path(log_dir), warnings=["no_data"])

    output_path = _build_output_path(log_dir)
    workbook = Workbook()

    sheets = {
        "tPP_AVG": workbook.active,
        "tPP_MAX": workbook.create_sheet("tPP_MAX"),
        "tPP_MIN": workbook.create_sheet("tPP_MIN"),
        "tBE": workbook.create_sheet("tBE"),
        "tSE": workbook.create_sheet("tSE"),
        "VBA": workbook.create_sheet("VBA"),
        "Sample_ID": workbook.create_sheet("Sample_ID"),
    }
    sheets["tPP_AVG"].title = "tPP_AVG"

    _init_sample_id_sheet(sheets["Sample_ID"])

    warnings: List[str] = []
    for index, port_data in enumerate(port_entries):
        _write_sample_id(sheets["Sample_ID"], port_data)
        _write_fail_sheet_if_needed(workbook, port_data)
        grid = _parse_lines(port_data.lines)
        if not grid:
            warnings.append(f"{port_data.name}: no parsed rows")
            continue

        _write_vba_sheet(sheets["VBA"], grid, _vba_start_col(index, cyc_type))

        summary = _calculate_summary(grid, cyc_type)
        if not summary:
            warnings.append(f"{port_data.name}: missing labels")
            continue

        _write_summary(sheets, index, port_data.name, summary, cyc_type)

    if "VBA" in workbook.sheetnames:
        workbook.remove(workbook["VBA"])

    _remove_empty_sheets(workbook)

    workbook.save(output_path)
    return SummaryResult(output_path=output_path, warnings=warnings)


def _build_output_path(log_dir: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return log_dir / f"auto_summary_{timestamp}.xlsx"


def _load_port_data(log_dir: Path) -> List[PortData]:
    port_map: Dict[str, List[str]] = {}
    log_files = list(log_dir.glob("*_log.txt"))
    for log_path in log_files:
        port_name = log_path.stem.replace("_log", "")
        with log_path.open("r", encoding="utf-8", newline="") as file_handle:
            lines = [line.rstrip("\n") for line in file_handle if line.strip()]
        if not lines:
            continue
        port_map.setdefault(port_name, []).extend(lines)

    return [PortData(name=port, lines=lines) for port, lines in port_map.items()]


def _parse_lines(lines: Iterable[str]) -> List[List[str]]:
    rows: List[List[str]] = []
    for line in lines:
        if ";" in line:
            parts = [part.strip() for part in line.split(";")]
        else:
            parts = [line.strip()]
        if parts:
            rows.append(parts)
    return rows


def _write_fail_sheet_if_needed(workbook: Workbook, port_data: PortData) -> None:
    tail_lines = [line.strip() for line in port_data.lines if line.strip()]
    if len(tail_lines) < 2:
        return
    if tail_lines[-2] == _PASS_LINE and tail_lines[-1] == _LOG_END_LINE:
        return

    start_index = None
    for idx in range(len(tail_lines) - 1, -1, -1):
        if "Result" in tail_lines[idx]:
            start_index = idx
            break
    if start_index is None:
        for idx in range(len(tail_lines) - 1, -1, -1):
            if "= = = = =" in tail_lines[idx]:
                start_index = idx
                break
    if start_index is None:
        return

    fail_lines = tail_lines[start_index:]
    if not fail_lines:
        return

    sheet_name = f"{port_data.name}_Fail"
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
    else:
        sheet = workbook.create_sheet(sheet_name)

    for row_index, line in enumerate(fail_lines, start=1):
        sheet.cell(row=row_index, column=1, value=_safe_text(line))


def _init_sample_id_sheet(sheet) -> None:
    sheet.cell(row=1, column=1, value="COM")
    sheet.cell(row=1, column=2, value="UID")


def _write_sample_id(sheet, port_data: PortData) -> None:
    uid = _extract_uid(port_data.lines)
    if not uid:
        return
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1, value=_extract_com_number(port_data.name))
    sheet.cell(row=row, column=2, value=uid)


def _vba_start_col(index: int, cyc_type: int) -> int:
    step = 18 if cyc_type in (0, 1, 3, 4) else 113
    return 1 + step * index


def _write_vba_sheet(sheet, grid: List[List[str]], start_col: int) -> None:
    for r_index, row in enumerate(grid, start=1):
        for c_index, value in enumerate(row, start=0):
            sheet.cell(
                row=r_index,
                column=start_col + c_index,
                value=_safe_text(value),
            )


def _safe_text(value: str) -> str:
    if value.startswith("="):
        return "'" + value
    return value


def _extract_uid(lines: List[str]) -> Optional[str]:
    for line in lines:
        match = _UID_PATTERN.search(line)
        if match:
            return match.group(1)
    return None


def _extract_com_number(port_name: str) -> str:
    return port_name.replace("COM", "")


def _find_label(grid: List[List[str]], label: str) -> Optional[Tuple[int, int]]:
    for r_index, row in enumerate(grid):
        for c_index, value in enumerate(row):
            if value == label:
                return r_index, c_index
    return None


def _numeric_values(row: List[str], start: int, count: int) -> List[float]:
    values: List[float] = []
    for offset in range(count):
        idx = start + offset
        if idx >= len(row):
            break
        try:
            values.append(float(row[idx]))
        except ValueError:
            continue
    return values


def _get_cell_value(grid: List[List[str]], row: int, col: int) -> Optional[float]:
    if row < 0 or row >= len(grid):
        return None
    if col < 0 or col >= len(grid[row]):
        return None
    try:
        return float(grid[row][col])
    except ValueError:
        return None


def _get_cell_int(grid: List[List[str]], row: int, col: int) -> Optional[int]:
    value = _get_cell_value(grid, row, col)
    if value is None:
        return None
    return int(value)


def _calculate_summary(grid: List[List[str]], cyc_type: int) -> Optional[Dict[str, List[str]]]:
    labels = {
        "avg": _find_label(grid, _LABEL_AVG),
        "max": _find_label(grid, _LABEL_MAX),
        "min": _find_label(grid, _LABEL_MAX),
        "tbe": _find_label(grid, _LABEL_TBE) or _find_label(grid, _LABEL_TBE_ALT),
        "tse": _find_label(grid, _LABEL_TSE) or _find_label(grid, _LABEL_TBE_ALT),
    }

    if not labels["avg"] or not labels["max"] or not labels["min"]:
        return None
    if cyc_type in (1, 4) and not labels["tse"]:
        return None
    if cyc_type not in (1, 4) and not labels["tbe"]:
        return None

    max_cyc = _max_first_column(grid)
    if max_cyc is None:
        return None
    v_range = int(max_cyc // 1000)

    avg_tpp: List[Optional[float]] = []
    max_tpp: List[Optional[float]] = []
    min_in_avg_tpp: List[Optional[float]] = []
    tbe: List[Optional[float]] = []
    tse: List[Optional[float]] = []
    cycle_time: List[Optional[int]] = []

    avg_row, avg_col = labels["avg"]
    max_row, max_col = labels["max"]
    min_row, min_col = labels["min"]
    tbe_pos = labels["tbe"]
    tse_pos = labels["tse"]
    if cyc_type in (1, 4) and tse_pos is None:
        return None
    if cyc_type not in (1, 4) and tbe_pos is None:
        return None

    if tse_pos is None:
        tse_pos = (0, 0)
    if tbe_pos is None:
        tbe_pos = (0, 0)

    for index in range(v_range + 1):
        row_avg = avg_row + 2 + index
        row_max = max_row + 2 + index
        row_min = min_row + 2 + index

        if cyc_type in (0, 3):
            avg_values = _numeric_values(grid[row_avg], avg_col + 1, 16)
            max_values = _numeric_values(grid[row_max], max_col + 1, 15)
            min_values = _numeric_values(grid[row_min], min_col + 1, 15)
            tbe_row = tbe_pos[0] + 2 + index
            tbe_values = _numeric_values(grid[tbe_row], tbe_pos[1] + 1, 15)
            cycle = _get_cell_int(grid, row_avg, tbe_pos[1])
            avg_tpp.append(_avg(avg_values))
            max_tpp.append(_max(max_values))
            min_in_avg_tpp.append(_avg(min_values))
            tbe.append(_max(tbe_values))
            cycle_time.append(cycle)

        elif cyc_type in (1, 4):
            avg_values = _numeric_values(grid[row_avg], avg_col + 1, 16)
            max_values = _numeric_values(grid[row_max], max_col + 1, 15)
            min_values = _numeric_values(grid[row_min], min_col + 1, 15)
            tse_row = tse_pos[0] + 2 + index
            tse_values = _numeric_values(grid[tse_row], tse_pos[1] + 1, 15)
            cycle = _get_cell_int(grid, row_avg, tse_pos[1])
            avg_tpp.append(_avg(avg_values))
            max_tpp.append(_max(max_values))
            min_in_avg_tpp.append(_avg(min_values))
            tse.append(_max(tse_values))
            cycle_time.append(cycle)

        else:
            avg_values = _numeric_values(grid[row_avg], avg_col + 1, 112)
            max_values = _numeric_values(grid[row_max], max_col + 1, 111)
            min_values = _numeric_values(grid[row_min], min_col + 1, 111)
            tbe_row = tbe_pos[0] + 2 + index
            tbe_values = _numeric_values(grid[tbe_row], tbe_pos[1] + 1, 111)
            cycle = _get_cell_int(grid, row_avg, tbe_pos[1])
            avg_tpp.append(_avg(avg_values))
            max_tpp.append(_max(max_values))
            min_in_avg_tpp.append(_avg(min_values))
            tbe.append(_max(tbe_values))
            cycle_time.append(cycle)

    return {
        "cycle_time": _prefix_int_values(cycle_time, "TPP_avg_"),
        "cycle_time_max": _prefix_int_values(cycle_time, "TPP_max_"),
        "cycle_time_min": _prefix_int_values(cycle_time, "TPP_MIN(AVG)_"),
        "cycle_time_tbe": _prefix_int_values(cycle_time, "tERS_"),
        "avg_tpp": _to_strings(avg_tpp),
        "max_tpp": _to_strings(max_tpp),
        "min_in_avg_tpp": _to_strings(min_in_avg_tpp),
        "tbe": _to_strings(tbe),
        "tse": _to_strings(tse),
    }


def _max_first_column(grid: List[List[str]]) -> Optional[float]:
    values: List[float] = []
    for row in grid:
        if not row:
            continue
        try:
            values.append(float(row[0]))
        except ValueError:
            continue
    return max(values) if values else None


def _avg(values: List[float]) -> Optional[float]:
    if not values:
        return None
    return sum(values) / len(values)


def _max(values: List[float]) -> Optional[float]:
    if not values:
        return None
    return max(values)


def _prefix_int_values(values: List[Optional[int]], prefix: str) -> List[str]:
    return [f"{prefix}{value}" if value is not None else "" for value in values]


def _to_strings(values: List[Optional[float]]) -> List[str]:
    return ["" if value is None else f"{value:.2f}" for value in values]


def _write_summary(
    sheets: Dict[str, Workbook],
    index: int,
    port_name: str,
    summary: Dict[str, List[str]],
    cyc_type: int,
) -> None:
    col = 2 + index

    _write_column(sheets["tPP_AVG"], col, port_name + "_AVG_TPP", summary["cycle_time"], summary["avg_tpp"])
    _write_column(sheets["tPP_MAX"], col, port_name + "_MAX_TPP", summary["cycle_time_max"], summary["max_tpp"])
    _write_column(sheets["tPP_MIN"], col, port_name + "_MIN(AVG)_TPP", summary["cycle_time_min"], summary["min_in_avg_tpp"])

    if summary["tbe"]:
        target_sheet = "tSE" if cyc_type == 3 else "tBE"
        target_title = port_name + ("_TSE" if cyc_type == 3 else "_TBE")
        _write_column(sheets[target_sheet], col, target_title, summary["cycle_time_tbe"], summary["tbe"])
    if summary["tse"]:
        _write_column(sheets["tSE"], col, port_name + "_TSE", summary["cycle_time_tbe"], summary["tse"])


def _write_column(sheet, col: int, title: str, x_values: List[str], y_values: List[str]) -> None:
    sheet.cell(row=1, column=1, value="Cycled Time")
    for idx, value in enumerate(x_values, start=2):
        sheet.cell(row=idx, column=1, value=value)

    sheet.cell(row=1, column=col, value=title)
    for idx, value in enumerate(y_values, start=2):
        sheet.cell(row=idx, column=col, value=value)


def _remove_empty_sheets(workbook: Workbook) -> None:
    for sheet_name in list(workbook.sheetnames):
        sheet = workbook[sheet_name]
        has_data = any(
            cell.value not in (None, "")
            for row in sheet.iter_rows()
            for cell in row
        )
        if not has_data and len(workbook.sheetnames) > 1:
            workbook.remove(sheet)
