"""
cz_corner_table.py  —  CZ Corner Summary 統計表格產生器  v2.0
==============================================================
將 CZ Corner Summary Excel 中每個 Split × 溫度 的表格區塊，
對 5 顆 sample 按 VCC 分組計算 AVG / STD / MAX / MIN，
並渲染為類似原始報告格式的 PNG 圖片。

支援檔案格式：
  - AAG106_ CZ_Corner_Summary.xlsx  (3 VCC/sample, 9 splits, 65 items)
  - 20240926_AAG095100_Coner CZ.xlsx (4 VCC/sample, 3 splits, 70 items, 3 sheets)

輸出：{out_dir}/{檔案}_{工作表}_{Split}_{溫度}.png

偵測邏輯：
  - Dark column (theme color 0/1) = 溫度區塊分隔線（vertical）
  - Dark row  in col A           = Split 分隔線（horizontal）
  - 每個 Split block 結構（行偏移）：
      +0  split name row  (AAG106: data col+1; AAG095100: label col)
      +1  temperature + sample numbers row
      +2  VCC values row
      +3+ characteristic items data rows
"""

import os
import re
import math
import time
import openpyxl
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib import rcParams

matplotlib.use('Agg')

# ══ 預設設定（可由 run_report.py 或命令列覆蓋）══════════════════════════════
DEFAULT_OUTPUT_DIR = 'charts/cz'
DEFAULT_FILES = [
    'AAG106_ CZ_Corner_Summary.xlsx',
    '20240926_AAG095100_Coner CZ.xlsx',
]

# ══ 單位字典（用於顯示）══════════════════════════════════════════════════════
UNIT_MAP = {
    'MHZ': 'MHz', 'KHZ': 'KHz',
    'NS':  'ns',  'PS':  'ps',  'US':  'µs',   'MS': 'ms',  'S': 's',
    'MA':  'mA',  'UA':  'µA',  'NA':  'nA',   'A':  'A',
    'V':   'V',   'MV':  'mV',
}

# ══ 圖表色彩主題 ════════════════════════════════════════════════════════════
COLORS = {
    'title_bg':    '#FFFF00',   # 黃色（標題欄，對應原始報告）
    'title_fg':    '#000000',
    'vcc_bgs':     ['#DDEEFF', '#FFEECC', '#EEF0EE'],   # 3 種 VCC 欄底色（可擴充）
    'vcc_fg':      '#1A202C',
    'stats_hdr_bg':'#BDD7EE',   # 統計標頭（淺藍）
    'stats_hdr_fg':'#1A202C',
    'item_bg':     '#FFFFFF',
    'item_fg':     '#1A202C',
    'odd_bg':      '#F5F5F5',
    'even_bg':     '#FFFFFF',
    'border':      '#AAAAAA',
    'spec_bg':     '#FFFF99',   # Spec 欄淡黃
}

# 全域字體設定（避免每張圖反覆設定 rcParams）
rcParams['font.family'] = 'sans-serif'


# ════════════════════════════════════════════════════════════════════════════
# 工具函式
# ════════════════════════════════════════════════════════════════════════════

def is_dark(cell) -> bool:
    """偵測是否為黑色分隔儲存格（theme color 0 or 1）"""
    fill = cell.fill
    if fill and fill.fgColor and fill.fgColor.type == 'theme':
        return str(fill.fgColor.value) in ('0', '1')
    return False


def parse_value_unit(raw) -> tuple:
    """
    從原始字串解析出 (float_value, unit_string)。
    例如：' 181.8MHZ  ' → (181.8, 'MHz')
         '-60.00NA  '  → (-60.0, 'nA')
         ' 0.000A   '  → (0.0,   'A')
    """
    if raw is None:
        return (None, '')
    s = str(raw).strip()
    m = re.match(r'^([+-]?\d+\.?\d*(?:E[+-]?\d+)?)\s*([A-Za-z]+)?$', s, re.I)
    if m:
        try:
            val  = float(m.group(1))
            unit_raw = (m.group(2) or '').upper()
            unit = UNIT_MAP.get(unit_raw, unit_raw)
            return (val, unit)
        except ValueError:
            pass
    return (None, '')


def fmt_num(v, digits=4) -> str:
    """格式化數字為易讀字串（最多 digits 位有效數字）"""
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return ''
    if v == 0:
        return '0'
    # 用 g 格式：移除尾零
    return f'{v:.{digits}g}'


def safe_name(s) -> str:
    """轉換為安全檔名片段"""
    return re.sub(r'[^\w\-]', '_', str(s)).strip('_')


# ════════════════════════════════════════════════════════════════════════════
# Excel 結構解析
# ════════════════════════════════════════════════════════════════════════════

def find_dark_col_positions(ws) -> list:
    """掃描前幾列，找出 dark column 的欄號（溫度區塊分隔）"""
    dark_cols = set()
    for r in range(1, min(ws.max_row + 1, 10)):
        for c in range(1, ws.max_column + 1):
            if is_dark(ws.cell(r, c)):
                dark_cols.add(c)
    return sorted(dark_cols)


def find_dark_row_positions(ws) -> list:
    """掃描 col A，找出 dark row 的列號（Split 分隔）"""
    return [r for r in range(1, ws.max_row + 1) if is_dark(ws.cell(r, 1))]


def parse_split_blocks(ws, dark_rows) -> list:
    """根據 dark rows 切分出各 Split block 的列範圍"""
    boundaries = [0] + dark_rows + [ws.max_row + 1]
    blocks = []
    for i in range(len(boundaries) - 1):
        start = boundaries[i] + 1
        end   = boundaries[i + 1] - 1
        if end >= start:
            blocks.append({'row_start': start, 'row_end': end})
    return blocks


def parse_temp_blocks(ws, dark_cols) -> list:
    """根據 dark cols 切分出各溫度 block 的欄範圍"""
    boundaries = [0] + dark_cols + [ws.max_column + 1]
    blocks = []
    for i in range(len(boundaries) - 1):
        start = boundaries[i] + 1
        end   = boundaries[i + 1] - 1
        if end >= start:
            blocks.append({'col_start': start, 'col_end': end})
    return blocks


def detect_split_name(ws, split_block, temp_blocks) -> str:
    """從第一個有效溫度欄塊的 +0 row 找出 split name"""
    rs = split_block['row_start']
    for tb in temp_blocks:
        cs, ce = tb['col_start'], tb['col_end']
        for col_try in [cs, cs + 1] if cs + 1 <= ce else [cs]:
            v = ws.cell(rs, col_try).value
            if v and str(v).strip():
                return str(v).strip()
    return 'Unknown'


def extract_block_stats(ws, split_block, temp_block, split_name=None) -> dict | None:
    """
    提取單一 Split × 溫度 區塊，計算每個特性項目在各 VCC 的統計值。

    返回 dict：
    {
      'split_name': str,
      'temp_name':  str,
      'vccs':       ['2.700V', '3.000V', ...],   # 排序後的唯一 VCC 清單
      'items':      ['FR-03 ...', ...],
      'units':      ['MHz', 'MHz', ...],           # 與 items 等長
      'stats':      [                              # 與 items 等長
          {'2.700V': {'AVG': x, 'STD': x, 'MAX': x, 'MIN': x}, ...},
          ...
      ]
    }
    或 None（空區塊）
    """
    rs  = split_block['row_start']
    re_ = split_block['row_end']
    cs  = temp_block['col_start']
    ce  = temp_block['col_end']

    # ── Split name ──
    if not split_name:
        for col_try in ([cs, cs + 1] if cs + 1 <= ce else [cs]):
            v = ws.cell(rs, col_try).value
            if v and str(v).strip():
                split_name = str(v).strip()
                break
    if not split_name:
        return None

    # ── Temperature（+1 row，label col）──
    r1 = rs + 1
    temp_val = ws.cell(r1, cs).value
    temp_name = str(temp_val).strip() if temp_val else 'Unknown'

    # ── Sample numbers（+1 row，data cols）──
    # col → sample_number
    sample_map = {}
    for c in range(cs + 1, ce + 1):
        v = ws.cell(r1, c).value
        if v is not None and str(v).strip().lstrip('-').isdigit():
            sample_map[c] = int(str(v).strip())

    # ── VCC 值（+2 row，data cols）──
    r2 = rs + 2
    # col → VCC string（strip 後）
    vcc_map = {}
    for c in range(cs + 1, ce + 1):
        v = ws.cell(r2, c).value
        if v and 'V' in str(v).upper():
            vcc_map[c] = str(v).strip()

    if not vcc_map:
        return None

    # 每個 VCC 對應的欄位清單（可能同一 VCC 有多個 sample）
    # vcc_string → [col, col, ...]
    vcc_cols: dict[str, list] = {}
    for c, vcc_s in vcc_map.items():
        vcc_cols.setdefault(vcc_s, []).append(c)

    # VCC 排序（按數值大小）
    def vcc_sort_key(vs):
        m = re.search(r'[\d.]+', vs)
        return float(m.group()) if m else 0
    sorted_vccs = sorted(vcc_cols.keys(), key=vcc_sort_key)

    # ── 資料列（+3 row 以後）──
    r_data_start = rs + 3
    items = []
    units_list = []
    stats_list = []

    for r in range(r_data_start, re_ + 1):
        item_name = ws.cell(r, cs).value
        if not item_name or not str(item_name).strip():
            continue

        item_unit = ''
        item_stats = {}

        for vcc_s in sorted_vccs:
            cols = vcc_cols[vcc_s]
            values = []
            for c in cols:
                raw = ws.cell(r, c).value
                val, unit = parse_value_unit(raw)
                if val is not None:
                    values.append(val)
                    if not item_unit and unit:
                        item_unit = unit

            if values:
                n = len(values)
                avg = sum(values) / n
                if n > 1:
                    variance = sum((x - avg) ** 2 for x in values) / (n - 1)
                    std = math.sqrt(variance)
                else:
                    std = 0.0
                item_stats[vcc_s] = {
                    'AVG': avg,
                    'STD': std,
                    'MAX': max(values),
                    'MIN': min(values),
                }
            else:
                item_stats[vcc_s] = {'AVG': None, 'STD': None, 'MAX': None, 'MIN': None}

        items.append(str(item_name).strip())
        units_list.append(item_unit)
        stats_list.append(item_stats)

    if not items:
        return None

    return {
        'split_name': split_name,
        'temp_name':  temp_name,
        'vccs':       sorted_vccs,
        'items':      items,
        'units':      units_list,
        'stats':      stats_list,
    }


# ════════════════════════════════════════════════════════════════════════════
# 表格渲染（自訂格線繪製，支援多層欄標頭）
# ════════════════════════════════════════════════════════════════════════════

STAT_COLS = ['AVG', 'STD', 'MAX', 'MIN']

# 欄寬比例（英寸）
COL_W_ITEM  = 2.2    # 項目名稱欄
COL_W_SPEC  = 0.55   # Spec 欄
COL_W_UNIT  = 0.55   # Unit 欄
COL_W_STAT  = 0.82   # 每個統計子欄（AVG/STD/MAX/MIN）

ROW_H_HDR   = 0.30   # 標頭列高（英寸）
ROW_H_DATA  = 0.22   # 資料列高（英寸）
FONT_HDR    = 7.5
FONT_DATA   = 6.5
DPI         = 130


def _draw_cell(ax, x, y, w, h,
               text='', bg='#FFFFFF', fg='#000000',
               fontsize=7, bold=False, halign='center', border_color='#AAAAAA'):
    """在 axes 座標系繪製單一儲存格（矩形 + 文字）"""
    # ponytail: Rectangle 比 FancyBboxPatch 輕量，批量畫表格更快
    rect = mpatches.Rectangle(
        (x, y), w, h,
        linewidth=0.4, edgecolor=border_color,
        facecolor=bg, clip_on=True,
    )
    ax.add_patch(rect)
    if text:
        ax.text(
            x + (w * (0.05 if halign == 'left' else 0.5)),
            y + h / 2,
            str(text),
            fontsize=fontsize,
            fontweight='bold' if bold else 'normal',
            color=fg,
            ha=halign, va='center',
            clip_on=True,
        )


def render_stats_png(data: dict, file_label: str, sheet: str, out_path: str):
    """將統計資料渲染為多層欄標頭的 PNG 表格圖片"""
    items      = data['items']
    vccs       = data['vccs']
    units_list = data['units']
    stats_list = data['stats']
    split_name = data['split_name']
    temp_name  = data['temp_name']

    n_items  = len(items)
    n_vccs   = len(vccs)
    n_stats  = len(STAT_COLS)   # 4

    # ── 欄位 x 位置計算 ──────────────────────────────────────────────────
    # 欄位結構：[Item] [Spec] [Unit] [VCC1: AVG STD MAX MIN] [VCC2: ...] ...
    col_xs    = [0]
    col_ws    = [COL_W_ITEM, COL_W_SPEC, COL_W_UNIT]
    for _ in vccs:
        col_ws += [COL_W_STAT] * n_stats
    for w in col_ws:
        col_xs.append(col_xs[-1] + w)
    total_w = col_xs[-1]

    # ── 列位 y 位置計算（從底部 0 往上）──────────────────────────────────
    # 我們用「從頂部往下」的邏輯，最後 flip y
    # row 0 = 標題列（Split + Temp）
    # row 1 = VCC 組標頭
    # row 2 = 統計欄標頭（AVG STD MAX MIN）
    # rows 3.. = 資料

    n_hdr_rows = 3
    total_h = (n_hdr_rows * ROW_H_HDR + n_items * ROW_H_DATA)
    fig_h   = total_h + 0.05   # 邊距

    fig, ax = plt.subplots(figsize=(total_w, fig_h))
    ax.set_xlim(0, total_w)
    ax.set_ylim(0, fig_h)
    ax.axis('off')
    fig.patch.set_facecolor('white')

    def row_y(row_idx) -> float:
        """row_idx 0 = 最頂列，往下遞增。回傳該列底部的 y 座標（from bottom）"""
        if row_idx < n_hdr_rows:
            top_of_row = fig_h - row_idx * ROW_H_HDR
        else:
            top_of_row = fig_h - n_hdr_rows * ROW_H_HDR - (row_idx - n_hdr_rows) * ROW_H_DATA
        return top_of_row - (ROW_H_HDR if row_idx < n_hdr_rows else ROW_H_DATA)

    def row_h(row_idx) -> float:
        return ROW_H_HDR if row_idx < n_hdr_rows else ROW_H_DATA

    # ── Row 0：標題列（Split + Temp）────────────────────────────────────
    ry = row_y(0)
    rh = row_h(0)
    title_text = f'{split_name}   {temp_name}'
    _draw_cell(ax, 0, ry, total_w, rh,
               text=title_text, bg=COLORS['title_bg'], fg=COLORS['title_fg'],
               fontsize=FONT_HDR + 1, bold=True, halign='left')

    # ── Row 1：VCC 組標頭 ────────────────────────────────────────────────
    ry = row_y(1)
    rh = row_h(1)
    # Item / Spec / Unit 欄（靜態標頭）
    static_cols = [
        (col_xs[0], col_ws[0], 'Item'),
        (col_xs[1], col_ws[1], 'Spec'),
        (col_xs[2], col_ws[2], 'Unit'),
    ]
    for sx, sw, st in static_cols:
        _draw_cell(ax, sx, ry, sw, rh,
                   text=st, bg=COLORS['stats_hdr_bg'], fg=COLORS['vcc_fg'],
                   fontsize=FONT_HDR, bold=True)

    # VCC 組大標頭（橫跨 4 個統計子欄）
    for vi, vcc in enumerate(vccs):
        x_start = col_xs[3 + vi * n_stats]
        group_w = COL_W_STAT * n_stats
        bg = COLORS['vcc_bgs'][vi % len(COLORS['vcc_bgs'])]
        _draw_cell(ax, x_start, ry, group_w, rh,
                   text=vcc, bg=bg, fg=COLORS['vcc_fg'],
                   fontsize=FONT_HDR, bold=True)

    # ── Row 2：統計子欄標頭 ──────────────────────────────────────────────
    ry = row_y(2)
    rh = row_h(2)
    for sx, sw, _ in static_cols:
        _draw_cell(ax, sx, ry, sw, rh,
                   text='', bg=COLORS['stats_hdr_bg'], fg=COLORS['vcc_fg'],
                   fontsize=FONT_HDR)

    for vi in range(n_vccs):
        bg = COLORS['vcc_bgs'][vi % len(COLORS['vcc_bgs'])]
        for si, stat in enumerate(STAT_COLS):
            cx = col_xs[3 + vi * n_stats + si]
            _draw_cell(ax, cx, ry, COL_W_STAT, rh,
                       text=stat, bg=bg, fg=COLORS['vcc_fg'],
                       fontsize=FONT_HDR, bold=True)

    # ── Data rows ────────────────────────────────────────────────────────
    for i, item in enumerate(items):
        ry = row_y(n_hdr_rows + i)
        rh = row_h(n_hdr_rows + i)
        bg_data = COLORS['odd_bg'] if i % 2 == 0 else COLORS['even_bg']

        # Item name
        _draw_cell(ax, col_xs[0], ry, col_ws[0], rh,
                   text=item, bg=bg_data, fg=COLORS['item_fg'],
                   fontsize=FONT_DATA, halign='left')

        # Spec（空白）
        _draw_cell(ax, col_xs[1], ry, col_ws[1], rh,
                   text='', bg=COLORS['spec_bg'], fg=COLORS['item_fg'],
                   fontsize=FONT_DATA)

        # Unit
        _draw_cell(ax, col_xs[2], ry, col_ws[2], rh,
                   text=units_list[i], bg=COLORS['spec_bg'], fg=COLORS['item_fg'],
                   fontsize=FONT_DATA)

        # Stats per VCC
        item_stats = stats_list[i]
        for vi, vcc in enumerate(vccs):
            vcc_stat = item_stats.get(vcc, {})
            for si, stat in enumerate(STAT_COLS):
                cx   = col_xs[3 + vi * n_stats + si]
                val  = vcc_stat.get(stat)
                text = fmt_num(val) if val is not None else ''
                _draw_cell(ax, cx, ry, COL_W_STAT, rh,
                           text=text, bg=bg_data, fg=COLORS['item_fg'],
                           fontsize=FONT_DATA)

    # ── 儲存 ────────────────────────────────────────────────────────────
    plt.savefig(out_path, dpi=DPI, facecolor='white', edgecolor='none')
    plt.close(fig)
    print(f'  OK {os.path.basename(out_path)}')


# ════════════════════════════════════════════════════════════════════════════
# 主流程
# ════════════════════════════════════════════════════════════════════════════

def process_worksheet(ws, file_label: str, sheet_name: str, out_dir: str,
                      meta_list: list):
    """處理單一工作表，產生所有 Split x 溫度 的統計圖，並將 metadata 加入 meta_list"""
    print(f'\n  Sheet: {sheet_name}')

    dark_rows    = find_dark_row_positions(ws)
    dark_cols    = find_dark_col_positions(ws)
    split_blocks = parse_split_blocks(ws, dark_rows)
    temp_blocks  = parse_temp_blocks(ws, dark_cols)

    print(f'    Splits: {len(split_blocks)},  Temp blocks: {len(temp_blocks)}')

    count = 0
    t0 = time.perf_counter()
    render_sec = 0.0
    for sb in split_blocks:
        sname = detect_split_name(ws, sb, temp_blocks)
        for tb in temp_blocks:
            data = extract_block_stats(ws, sb, tb, split_name=sname)
            if data is None:
                continue

            split_s = safe_name(data['split_name'])
            temp_s  = safe_name(data['temp_name'])
            fname   = f'{file_label}_{sheet_name}_{split_s}_{temp_s}.png'
            out_path = os.path.join(out_dir, fname)

            t_render = time.perf_counter()
            render_stats_png(data, file_label, sheet_name, out_path)
            render_sec += time.perf_counter() - t_render

            # 記錄 metadata（product = sheet_name；temp / split 來自資料，不 hard code）
            meta_list.append({
                'path':    os.path.abspath(out_path),
                'product': sheet_name,
                'temp':    data['temp_name'],
                'split':   data['split_name'],
                'file':    file_label,
            })
            count += 1

    total_sec = time.perf_counter() - t0
    print(f'    -> {count} images written')
    print(f'    -> timing: total={total_sec:.2f}s, render={render_sec:.2f}s, other={max(total_sec-render_sec, 0):.2f}s')


def process_file(xlsx_path: str, out_dir: str, meta_list: list):
    """處理單一 Excel 檔案（可多 sheets）"""
    file_label = safe_name(os.path.splitext(os.path.basename(xlsx_path))[0])[:25]
    print(f'\n{"=" * 60}')
    print(f'File: {os.path.basename(xlsx_path)}')

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        process_worksheet(ws, file_label, sheet_name, out_dir, meta_list)


def run(files=None, out_dir=None):
    """
    主入口：可由 run_report.py 呼叫，傳入自訂的 files 清單與輸出目錄。
    若未傳入則使用 DEFAULT_FILES / DEFAULT_OUTPUT_DIR。
    回傳：meta_list（每張圖的 product / temp / split / path）
    """
    import json

    if files is None:
        files = DEFAULT_FILES
    if out_dir is None:
        out_dir = DEFAULT_OUTPUT_DIR

    os.makedirs(out_dir, exist_ok=True)
    meta_list = []

    for f in files:
        if os.path.exists(f):
            process_file(f, out_dir, meta_list)
        else:
            print(f'WARNING: file not found: {f}')

    # 寫出 JSON 索引（供 gen_eng_report.py 讀取，動態插圖）
    index_path = os.path.join(out_dir, 'cz_index.json')
    with open(index_path, 'w', encoding='utf-8') as fp:
        json.dump(meta_list, fp, ensure_ascii=False, indent=2)
    print(f'\nIndex written: {index_path}  ({len(meta_list)} entries)')
    print(f'Done! All images saved to: {out_dir}/')
    return meta_list


if __name__ == '__main__':
    run()
