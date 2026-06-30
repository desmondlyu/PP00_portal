"""
cp_summary_report.py
====================
CP Summary Report Generator  —  single entry point.
可獨立執行，也可由 run_report.py import 並覆蓋 CONFIG。

Outputs (all under OUTPUT_DIR/):
  • 4 × PNG summary tables   summary_{GCCD|AACD|NHPH|NLPL}.png
  • 20 × PNG bar charts      {station}_{group}.png (grouped by split, top-5 fail categories)
  • 1 × Markdown fragment    _chart_sections.md

Standalone usage:
    python cp_summary_report.py

Module usage (from run_report.py):
    import cp_summary_report as cpr
    cpr.WORKBOOK   = 'FAG102_CP_Summary.xlsx'
    cpr.PRODUCT_NO = 'FAG102B'
    ...
    charts_map = cpr.build_all()   # → {station: {group_key: [fc1,fc2,fc3,fc4,fc5]}}
"""

import os
import re
import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.ticker as mticker
import numpy as np
from collections import defaultdict

# ═══════════════════════════════════════════════════════════════════════════════
#  RUNTIME VARIABLES  ← 全部由 run_report.py (UI) 動態注入，此處僅為空值宣告
# ═══════════════════════════════════════════════════════════════════════════════
PRODUCT_NO = ''    # 用於圖表標題，由 run_report.py 設定
WORKBOOK   = ''    # CP Summary Excel 路徑，由 run_report.py 設定
OUTPUT_DIR = 'charts'
STATIONS   = []    # 由 run_report.py 設定
PASS_BINS  = {'01', '02', '03', '04', '05', '07'}   # 固定業務規則

# Group definitions: split key (Excel) → short display label, ordered SS→POR→FF
GROUPS = {
    'GC CD': {
        'key': 'GCCD',
        'ordered': [
            ('GC CD -3 std',   '-3σ'),
            ('GC CD -1.5 std', '-1.5σ'),
            ('POR',            'POR'),
            ('GC CD +1.5 std', '+1.5σ'),
            ('GC CD +3 std',   '+3σ'),
        ],
    },
    'AA CD': {
        'key': 'AACD',
        'ordered': [
            ('AA CD -3 std',   '-3σ'),
            ('AA CD -1.5 std', '-1.5σ'),
            ('POR',            'POR'),
            ('AA CD +1.5 std', '+1.5σ'),
            ('AA CD +3 std',   '+3σ'),
        ],
    },
    'NH/PH': {
        'key': 'NHPH',
        'ordered': [
            ('NH/PH SS +3std',    'SS+3σ'),
            ('PH /NH SS +1.5std', 'SS+1.5σ'),
            ('POR',               'POR'),
            ('PH /NH FF -1.5std', 'FF-1.5σ'),
            ('NH/PH FF -3std',    'FF-3σ'),
        ],
    },
    'NL/PL': {
        'key': 'NLPL',
        # Two SS+3σ and two FF-3σ from different lots; Lot column distinguishes them
        'ordered': [
            ('NL /PL SS +3std', 'SS+3σ'),
            ('NL/PL SS+3std',   'SS+3σ'),
            ('POR',             'POR'),
            ('NL/PL FF -3std',  'FF-3σ'),
            ('NL /PL FF-3std',  'FF-3σ'),
        ],
    },
}

# ── Bar chart colors ──────────────────────────────────────────────────────────
BAR_COLOR = '#4472C4'   # Excel blue  (non-POR)
POR_COLOR = '#ED7D31'   # orange      (POR reference bar)

# ── Summary table colors ──────────────────────────────────────────────────────
C_HDR = '#548235'   # dark green header row
C_STA = '#E2EFDA'   # light green station-name cell
C_POR = '#FCE4D6'   # salmon  POR row
C_POS = '#DEEAF1'   # light blue  positive / FF condition
C_NEG = '#FFFFFF'   # white       negative / SS condition


def normalize_split(split_name):
    """
    Normalize split text for tolerant matching.
    Handles spacing variants like:
      NL/PL FF -3std == NL /PL FF-3 std
    """
    if split_name is None:
        return ''

    s = str(split_name).strip().upper().replace('／', '/')
    if not s:
        return ''
    if s == 'POR':
        return 'POR'

    # collapse all spaces to absorb user input variants
    s = re.sub(r'\s+', '', s)
    s = s.replace('STD.', 'STD')

    # tolerant domain normalization: SS is slow(+), FF is fast(-)
    s = s.replace('SS-', 'SS+').replace('FF+', 'FF-')
    return s


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA LOADING
# ═══════════════════════════════════════════════════════════════════════════════
def load_station(ws):
    """
    Read one Excel sheet (= one test station).
    Returns (sd, fail_cols):
      sd         – dict {split_name: {lots, pr_ylds, wf_ylds, avg_pr, avg_wf, avg_fail}}
      fail_cols  – ordered list of fail bin column names
    """
    headers   = [c.value for c in ws[1]]
    idx       = {h: i for i, h in enumerate(headers)}
    ak_start  = idx.get('01')
    if ak_start is None:
        raise ValueError(f"Cannot find bin column '01' in sheet '{ws.title}'")

    fail_cols = [c for c in headers[ak_start:]
                 if c is not None and str(c).strip() not in PASS_BINS]

    lot_key   = next((k for k in idx if str(k).lower().replace(' ', '_') in ('lot_no', 'lot')), None)
    split_key = next((k for k in idx if str(k).lower() == 'split'), None)
    pr_key    = next((k for k in idx if str(k).strip() == 'Pr. Yld.'), None)
    wf_key    = next((k for k in idx if str(k).strip() == 'Wf. Yld.'), None)

    sd = defaultdict(lambda: {
        'lots': set(), 'pr_ylds': [], 'wf_ylds': [],
        'fail_sum': defaultdict(float), 'n': 0,
    })

    for row in ws.iter_rows(min_row=2, values_only=True):
        sp = row[idx[split_key]] if split_key else None
        if sp is None:
            continue
        sp = normalize_split(sp)
        if not sp:
            continue
        d  = sd[sp]
        d['n'] += 1
        if lot_key:
            lot = row[idx[lot_key]]
            if lot:
                d['lots'].add(str(lot).strip())
        pr = row[idx[pr_key]] if pr_key else None
        wf = row[idx[wf_key]] if wf_key else None
        if pr: d['pr_ylds'].append(float(pr))
        if wf: d['wf_ylds'].append(float(wf))
        for fc in fail_cols:
            v = row[idx[fc]] if fc in idx else None
            if v:
                try:
                    d['fail_sum'][fc] += float(v)
                except (TypeError, ValueError):
                    pass

    for d in sd.values():
        n = d['n'] or 1
        d['avg_pr']   = sum(d['pr_ylds']) / len(d['pr_ylds']) if d['pr_ylds'] else 0
        d['avg_wf']   = sum(d['wf_ylds']) / len(d['wf_ylds']) if d['wf_ylds'] else 0
        d['avg_fail'] = {k: v / n for k, v in d['fail_sum'].items()}

    return sd, fail_cols


def group_top_n(sd, ordered, fail_cols, n):
    """
    Return top-n fail categories ranked by average fail % across
    non-POR splits present in the group.
    """
    agg = defaultdict(list)
    for raw_sp, _ in ordered:
        sp = normalize_split(raw_sp)
        if sp == 'POR' or sp not in sd:
            continue
        for fc in fail_cols:
            agg[fc].append(sd[sp]['avg_fail'].get(fc, 0))
    mean = {k: sum(v) / len(v) for k, v in agg.items() if any(v)}
    return [k for k, _ in sorted(mean.items(), key=lambda x: -x[1])[:n]]


# ═══════════════════════════════════════════════════════════════════════════════
#  PART 1 — SUMMARY TABLE IMAGES  (4 PNG, one per group)
# ═══════════════════════════════════════════════════════════════════════════════
def fmt_lot(lots_set):
    """Always show complete lot numbers, joined by ' / ' if multiple."""
    return ' / '.join(sorted(lots_set))


def row_bg(short_label):
    """Row background colour based on condition type."""
    lbl = short_label.upper()
    if lbl == 'POR':
        return C_POR
    if 'SS' in lbl or lbl.startswith('-'):
        return C_NEG   # slow / tight → white
    return C_POS       # fast / loose → light blue


def _generate_summary_image(grp_name, grp_info, all_data, outpath):
    station_blocks = []
    for sname in STATIONS:
        if sname not in all_data:
            continue
        sd, fail_cols = all_data[sname]
        top5 = group_top_n(sd, grp_info['ordered'], fail_cols, n=5)
        if not top5:
            continue

        col_labels = ['Process', 'Lot', 'Split', 'Pr.Yld.(%)', 'Wf.Yld.(%)'] + top5
        rows, cell_colors = [], []
        row_idx = 0

        for raw_sp, short_lbl in grp_info['ordered']:
            sp = normalize_split(raw_sp)
            if sp not in sd:
                continue
            d    = sd[sp]
            proc = sname if row_idx == 0 else ''
            bg   = row_bg(short_lbl)

            row = (
                [proc, fmt_lot(d['lots']), short_lbl,
                 f"{d['avg_pr']:.2f}", f"{d['avg_wf']:.2f}"]
                + [f"{d['avg_fail'].get(fc, 0):.3f}" for fc in top5]
            )
            rows.append(row)
            cell_colors.append([C_STA if (j == 0 and proc) else bg
                                 for j in range(len(row))])
            row_idx += 1

        if rows:
            station_blocks.append({'col_labels': col_labels,
                                    'rows': rows, 'cell_colors': cell_colors})

    if not station_blocks:
        return

    RAW_CW = [0.08, 0.30, 0.09, 0.09, 0.09] + [0.075] * 5
    total  = sum(RAW_CW)
    CW     = [w / total for w in RAW_CW]

    row_heights = [len(b['rows']) + 1 for b in station_blocks]
    fig_height  = sum(row_heights) * 0.38 + 0.55

    fig = plt.figure(figsize=(14, fig_height), facecolor='white')
    fig.text(0.5, 0.995,
             f'{grp_name} Split — Yield & Top-5 Fail Category Summary',
             ha='center', va='top', fontsize=13, fontweight='bold', color='#1F4E79')

    gs = gridspec.GridSpec(len(station_blocks), 1, figure=fig,
                           height_ratios=row_heights,
                           hspace=0.06, top=0.97, bottom=0.01,
                           left=0.005, right=0.995)

    for i, block in enumerate(station_blocks):
        ax = fig.add_subplot(gs[i])
        ax.axis('off')
        n_col  = len(block['col_labels'])
        n_rows = len(block['rows'])

        table = ax.table(
            cellText    = block['rows'],
            colLabels   = block['col_labels'],
            colWidths   = CW[:n_col],
            bbox        = [0, 0, 1, 1],
            cellLoc     = 'center',
            cellColours = block['cell_colors'],
        )
        table.auto_set_font_size(False)

        for j in range(n_col):                         # header
            cell = table[0, j]
            cell.set_facecolor(C_HDR)
            cell.set_text_props(color='white', fontweight='bold', fontsize=9)
            cell.set_linewidth(0.5)

        for r in range(n_rows):                        # data rows
            for j in range(n_col):
                tbl_cell = table[r + 1, j]
                tbl_cell.set_fontsize(8.5)
                tbl_cell.set_linewidth(0.3)
                if j >= 3:
                    tbl_cell.set_text_props(ha='right')
                if j == 0 and block['rows'][r][0]:
                    tbl_cell.set_text_props(fontweight='bold', ha='center',
                                            color='#375623')
                if block['rows'][r][2] == 'POR':
                    tbl_cell.set_text_props(fontweight='bold', color='#843C0C',
                                            ha='right' if j >= 3 else 'center')

    plt.savefig(outpath, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
    print(f'  ✓ {outpath}')


def generate_summary_tables(all_data):
    print('\n[1/2] Generating summary table images …')
    for grp_name, grp_info in GROUPS.items():
        outpath = os.path.join(OUTPUT_DIR, f"summary_{grp_info['key']}.png")
        _generate_summary_image(grp_name, grp_info, all_data, outpath)
    print('  → 4 summary tables done.')


# ═══════════════════════════════════════════════════════════════════════════════
#  PART 2 — BAR CHARTS  (60 PNG + markdown fragment)
# ═══════════════════════════════════════════════════════════════════════════════
def _draw_bar_chart(sd, ordered, top_fcs, grp_name, station, outpath):
    """
    Grouped bar chart: X-axis = splits, bars grouped by fail category.
    
    X-axis: splits (e.g., -3σ, -1.5σ, POR, +1.5σ, +3σ)
    Y-axis: fail category %
    Grouped bars: top_fcs (top 5 fail categories for this station/group)
    """
    split_labels = []
    fail_cat_data = {fc: [] for fc in top_fcs}
    
    for raw_sp, lbl in ordered:
        sp = normalize_split(raw_sp)
        if sp not in sd:
            continue
        split_labels.append(lbl)
        for fc in top_fcs:
            fail_cat_data[fc].append(sd[sp]['avg_fail'].get(fc, 0))
    
    n_splits = len(split_labels)
    if n_splits == 0:
        return
    
    # Colors for each fail category (cycling palette)
    colors = ['#4472C4', '#ED7D31', '#A5A5A5', '#FFC000', '#5B9BD5',
              '#70AD47', '#C55A11', '#666666', '#1F497D', '#4BACC6']
    fc_colors = {fc: colors[i % len(colors)] for i, fc in enumerate(top_fcs)}
    
    fig, ax = plt.subplots(figsize=(10, 5))
    x = np.arange(n_splits)
    bar_width = 0.15
    
    for i, fc in enumerate(top_fcs):
        offset = (i - len(top_fcs)/2 + 0.5) * bar_width
        bars = ax.bar(x + offset, fail_cat_data[fc], bar_width,
                      label=str(fc), color=fc_colors[fc], alpha=0.8)
        # Add value labels on bars
        for bar, val in zip(bars, fail_cat_data[fc]):
            if val > 0:
                ax.text(bar.get_x() + bar.get_width()/2, val, f'{val:.2f}',
                       ha='center', va='bottom', fontsize=7)
    
    ax.set_xticks(x)
    ax.set_xticklabels(split_labels, fontsize=9)
    ax.set_ylabel('Fail Cat. %', fontsize=9)
    ax.set_title(f'{PRODUCT_NO}  {station}  {grp_name}\nTop-5 Fail Categories across Splits',
                fontsize=10, fontweight='bold', pad=8)
    ymax = max(max(vals) for vals in fail_cat_data.values()) * 1.2 if any(fail_cat_data.values()) else 0.1
    ax.set_ylim(0, ymax)
    ax.yaxis.grid(True, linestyle='--', linewidth=0.5, alpha=0.6, zorder=0)
    ax.set_axisbelow(True)
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.legend(loc='upper right', fontsize=8, framealpha=0.9)
    
    plt.tight_layout(pad=1.0)
    plt.savefig(outpath, dpi=130, bbox_inches='tight')
    plt.close()


def generate_bar_charts(all_data):
    """
    新結構：每個 (Station, Group) 一張圖，包含 top-5 fail categories。
    回傳 charts_map：{station: {group_key: [fc1, fc2, fc3, fc4, fc5]}}
    """
    print('\n[2/2] Generating grouped bar charts …')
    md_lines   = []
    total      = 0
    charts_map = {}

    for sname in STATIONS:
        if sname not in all_data:
            continue
        sd, fail_cols = all_data[sname]
        md_lines.append(f'\n### {sname}\n')
        charts_map[sname] = {}

        for grp_name, gcfg in GROUPS.items():
            # Get top-5 fail categories for this station/group
            top5  = group_top_n(sd, gcfg['ordered'], fail_cols, n=5)
            if not top5:
                continue
            
            charts_map[sname][gcfg['key']] = top5
            fname = os.path.join(OUTPUT_DIR, f'{sname}_{gcfg["key"]}.png')
            _draw_bar_chart(sd, gcfg['ordered'], top5, grp_name, sname, fname)
            total += 1
            
            # Markdown: single image per group
            md_lines.append(f'#### {grp_name}\n')
            md_lines.append(f'![{sname} {grp_name}]({fname})\n')

    with open(os.path.join(OUTPUT_DIR, '_chart_sections.md'), 'w', encoding='utf-8') as f:
        f.write('\n'.join(md_lines))

    print(f'  → {total} grouped bar charts done.')
    print(f'  → Markdown fragment → {OUTPUT_DIR}/_chart_sections.md')
    return charts_map


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def build_all():
    """
    完整執行 CP Summary 圖表產生流程。
    供 run_report.py 呼叫，回傳 charts_map：
        {station: {group_key: [failcode1, failcode2, failcode3]}}
    """
    if os.path.exists(OUTPUT_DIR):
        for f in os.listdir(OUTPUT_DIR):
            try:
                os.remove(os.path.join(OUTPUT_DIR, f))
            except Exception:
                pass
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print('Loading Excel …')
    wb       = openpyxl.load_workbook(WORKBOOK, data_only=True)
    all_data = {}
    for sname in STATIONS:
        if sname in wb.sheetnames:
            all_data[sname] = load_station(wb[sname])
            print(f'  ✓ {sname}: {len(all_data[sname][0])} splits')

    generate_summary_tables(all_data)
    charts_map = generate_bar_charts(all_data)

    print(f'\n✅ Charts done.')
    print(f'   Summary tables : {OUTPUT_DIR}/summary_{{GCCD,AACD,NHPH,NLPL}}.png')
    print(f'   Bar charts     : 20 grouped PNG files in {OUTPUT_DIR}/ (5 stations × 4 groups)')
    return charts_map


def main():
    """獨立執行入口（已停用，請使用 run_report.py）。"""
    print('⚠️  請透過 run_report.py (UI) 執行，此檔案不支援獨立執行。')
    print('    python run_report.py')


if __name__ == '__main__':
    main()
